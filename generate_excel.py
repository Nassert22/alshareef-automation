import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

filename = "alshareef-automation.xlsx"

if os.path.exists(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["الرقم", "الوصف", "التاريخ"])

# نضيف صف جديد (كمثال بسيط)
ws.append([
    ws.max_row, 
    "تم التعديل من Codex وتم التحديث تلقائياً", 
    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
])

wb.save(filename)
print("✅ تم تحديث ملف Excel بنجاح.")
